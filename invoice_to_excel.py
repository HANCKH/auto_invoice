"""
invoice_to_excel.py
====================

Parse electronic VAT invoice PDFs and generate Excel import rows from a template.

Supports both single-file mode and batch mode (scan a folder). In batch mode,
results are grouped by vendor in subfolders named by vendor, and output files are
named as "<vendor>_<amount>.xlsx". PDF files can be copied/renamed to the
same pattern when requested.

Requirements:
    pip install pdfplumber openpyxl

"""

from __future__ import annotations

import argparse
import os
import re
import shutil
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Dict, Optional

import pdfplumber
import openpyxl
from openpyxl.utils import get_column_letter


def _clean_text(text: str) -> str:
    """Normalize OCR-like spaces and symbols in extracted text."""
    if not text:
        return ""
    text = text.replace("\u00a0", " ")
    text = text.replace("¥", "¥")
    text = re.sub(r"[\t\r]+", " ", text)
    return text


def _to_float(value: Optional[str]) -> float:
    if not value:
        return 0.0
    cleaned = str(value).replace(",", "").replace("，", "")
    try:
        return float(cleaned)
    except ValueError:
        match = re.search(r"(-?\d+(?:\.\d+)?)", cleaned)
        return float(match.group(1)) if match else 0.0


def _to_int(value: Optional[str]) -> int:
    if not value:
        return 1
    try:
        return int(float(str(value).replace(",", "")))
    except ValueError:
        return 1


def _normalize_unit(value: Optional[str]) -> str:
    unit = (value or "").strip()
    if not unit:
        return "个"
    # If unit is English-like token, normalize to Chinese default unit.
    if re.search(r"[A-Za-z]", unit):
        return "个"
    return unit


def _normalize_money(value: Optional[str]) -> str:
    if not value:
        return ""
    num = _to_float(value)
    # keep 2 decimals, then trim trailing zeros and dots to match sample names
    return (f"{num:.6f}".rstrip("0").rstrip(".")).rstrip(".")


def _safe_filename(name: str, max_len: int = 80) -> str:
    name = name.strip() if name else ""
    if not name:
        return "未识别"
    name = re.sub(r"[\\/:*?\"<>|]", "_", name)
    name = re.sub(r"\s+", " ", name)
    return name[:max_len] if len(name) > max_len else name


def _extract_money_by_keywords(text: str) -> tuple[Optional[str], Optional[str], Optional[str]]:
    amount = tax = total = None

    # Most e-invoices in samples expose amount/tax in "合计 ¥x ¥y"
    sum_line = re.search(
        r"合\s*计\s*[¥￥]?\s*([0-9]+(?:[.,][0-9]{1,6})?)\s*[¥￥]?\s*([0-9]+(?:[.,][0-9]{1,6})?)",
        text,
    )
    if sum_line:
        amount = sum_line.group(1).replace(",", ".")
        tax = sum_line.group(2).replace(",", ".")

    # "价税合计（小写）¥xxx.xx" is usually the most stable total source.
    total_small = re.search(
        r"[（(]\s*小\s*写\s*[)）]\s*[¥￥]?\s*([0-9]+(?:[.,][0-9]{1,6})?)",
        text,
    )
    if total_small:
        total = total_small.group(1).replace(",", ".")

    if not all([amount, tax, total]):
        # fallback: keyword form where values may be separated by spaces/newlines
        amount_m = re.search(r"金\s*额\s*[:：]?\s*[¥￥]?\s*([0-9]+(?:[.,][0-9]{1,6})?)", text)
        tax_m = re.search(r"税\s*额\s*[:：]?\s*[¥￥]?\s*([0-9]+(?:[.,][0-9]{1,6})?)", text)
        total_m = re.search(r"价\s*税\s*合\s*计[\s\S]{0,20}[¥￥]\s*([0-9]+(?:[.,][0-9]{1,6})?)", text)
        if amount_m and not amount:
            amount = amount_m.group(1).replace(",", ".")
        if tax_m and not tax:
            tax = tax_m.group(1).replace(",", ".")
        if total_m and not total:
            total = total_m.group(1).replace(",", ".")

    if not all([amount, tax, total]):
        # fallback: extract numbers strictly from currency-marked lines only.
        pair = re.search(r"[¥￥]\s*([0-9]+(?:[.,][0-9]{1,6})?)\s+[¥￥]\s*([0-9]+(?:[.,][0-9]{1,6})?)", text)
        if pair:
            amount = amount or pair.group(1).replace(",", ".")
            tax = tax or pair.group(2).replace(",", ".")

        currency_nums = re.findall(r"[¥￥]\s*([0-9]+(?:[.,][0-9]{1,6})?)", text)
        if currency_nums and not total:
            total = currency_nums[-1].replace(",", ".")

    return amount, tax, total


def _extract_product_line(text: str) -> Dict[str, Optional[str]]:
    data = {"product_line": None, "specification": None}
    lines = [ln.strip() for ln in _clean_text(text).splitlines() if ln.strip()]

    # pick a line likely containing product detail
    product_line = None
    for i, line in enumerate(lines):
        if line.startswith("*"):
            product_line = line
            break
        if "*" in line and not line.startswith("\u002a"):
            product_line = line
            break

    if not product_line:
        for line in lines:
            if "商品" in line and any(ch.isdigit() for ch in line) and any(k in line for k in ["个", "片", "件", "套", "台"]):
                product_line = line
                break

    if product_line:
        parts = product_line.split()
        if parts:
            data["product_line"] = parts[0]
            spec_tokens = parts[1:]
            unit_set = {"个", "片", "件", "盒", "瓶", "米", "套", "袋", "台", "只", "块", "箱", "批"}

            def _is_num(token: str) -> bool:
                return bool(re.fullmatch(r"\d+(?:\.\d+)?", token))

            def _is_rate(token: str) -> bool:
                return bool(re.fullmatch(r"\d+(?:\.\d+)?%", token))

            # Strip financial tail from right:
            # tax -> rate -> amount -> unit_price -> qty(optional) -> unit(optional)
            if spec_tokens and _is_num(spec_tokens[-1]):
                spec_tokens.pop()
            if spec_tokens and _is_rate(spec_tokens[-1]):
                spec_tokens.pop()
            if spec_tokens and _is_num(spec_tokens[-1]):
                spec_tokens.pop()
            if spec_tokens and _is_num(spec_tokens[-1]):
                spec_tokens.pop()
            if len(spec_tokens) >= 2 and _is_num(spec_tokens[-1]) and spec_tokens[-2] in unit_set:
                spec_tokens.pop()
            if spec_tokens and spec_tokens[-1] in unit_set:
                spec_tokens.pop()

            data["specification"] = " ".join(spec_tokens).strip()
        else:
            data["product_line"] = product_line
            data["specification"] = ""

    return data


def _extract_invoice_items(text: str) -> list[Dict[str, Optional[str]]]:
    """Extract all invoice item lines (one row per '*...' line)."""
    items: list[Dict[str, Optional[str]]] = []
    lines = [ln.strip() for ln in _clean_text(text).splitlines() if ln.strip()]
    units = r"(个|片|件|盒|瓶|米|套|袋|台|只|块|箱|批|pcs|PCS|Pc|PC|Pcs)"

    # Typical item tail:
    # "<unit> <qty> <unit_price> <amount> <tax_rate> <tax>"
    pat_full = re.compile(
        rf"^(?P<left>.+?)\s+{units}\s+(?P<qty>\d+(?:\.\d+)?)\s+"
        r"(?P<unit_price>\d+(?:\.\d+)?)\s+(?P<amount>-?\d+(?:\.\d+)?)\s+"
        r"(?P<rate>\d+(?:\.\d+)?%)\s+(?P<tax>-?\d+(?:\.\d+)?)$"
    )

    # Some layouts may omit quantity field in extracted text.
    pat_no_qty = re.compile(
        rf"^(?P<left>.+?)\s+{units}\s+"
        r"(?P<unit_price>\d+(?:\.\d+)?)\s+(?P<amount>-?\d+(?:\.\d+)?)\s+"
        r"(?P<rate>\d+(?:\.\d+)?%)\s+(?P<tax>-?\d+(?:\.\d+)?)$"
    )
    pat_full_any = re.compile(
        rf"(?P<left>.+?)\s+{units}\s+(?P<qty>\d+(?:\.\d+)?)\s+"
        r"(?P<unit_price>\d+(?:\.\d+)?)\s+(?P<amount>-?\d+(?:\.\d+)?)\s+"
        r"(?P<rate>\d+(?:\.\d+)?%)\s+(?P<tax>-?\d+(?:\.\d+)?)(?:\s|$)"
    )
    pat_no_qty_any = re.compile(
        rf"(?P<left>.+?)\s+{units}\s+"
        r"(?P<unit_price>\d+(?:\.\d+)?)\s+(?P<amount>-?\d+(?:\.\d+)?)\s+"
        r"(?P<rate>\d+(?:\.\d+)?%)\s+(?P<tax>-?\d+(?:\.\d+)?)(?:\s|$)"
    )
    # Discount/reduction row like: "*... -60.18 13% -7.82"
    pat_discount = re.compile(
        r"^(?P<left>.+?)\s+(?P<amount>-\d+(?:\.\d+)?)\s+"
        r"(?P<rate>\d+(?:\.\d+)?%)\s+(?P<tax>-\d+(?:\.\d+)?)$"
    )
    # No explicit unit in extracted text: "<left> <qty> <unit_price> <amount> <rate> <tax>"
    pat_no_unit = re.compile(
        r"^(?P<left>.+?)\s+(?P<qty>\d+(?:\.\d+)?)\s+"
        r"(?P<unit_price>\d+(?:\.\d+)?)\s+(?P<amount>-?\d+(?:\.\d+)?)\s+"
        r"(?P<rate>\d+(?:\.\d+)?%)\s+(?P<tax>-?\d+(?:\.\d+)?)$"
    )
    pat_no_unit_any = re.compile(
        r"(?P<left>.+?)\s+(?P<qty>\d+(?:\.\d+)?)\s+"
        r"(?P<unit_price>\d+(?:\.\d+)?)\s+(?P<amount>-?\d+(?:\.\d+)?)\s+"
        r"(?P<rate>\d+(?:\.\d+)?%)\s+(?P<tax>-?\d+(?:\.\d+)?)(?:\s|$)"
    )
    pat_concat_qty_price = re.compile(
        r"^(?P<left>.+?)\s+(?P<qprice>\d+\.\d+)\s+"
        r"(?P<amount>-?\d+(?:\.\d+)?)\s+(?P<rate>\d+(?:\.\d+)?%)\s+(?P<tax>-?\d+(?:\.\d+)?)$"
    )

    stop_prefixes = (
        "合 计",
        "价税合计",
        "备",
        "注",
        "开票人",
        "销方开户银行",
        "购方开户银行",
        "下载次数",
        "项目名称",
    )

    def _clean_suffix_text(suffix_text: str) -> str:
        """Keep descriptive continuation text, drop trailing monetary/tax tokens."""
        tokens = re.split(r"\s+", suffix_text.strip())
        kept: list[str] = []
        for tk in tokens:
            if not tk:
                continue
            # Ignore isolated numeric continuation fragments such as standalone "1".
            if re.fullmatch(r"\d+", tk):
                continue
            if "¥" in tk or "￥" in tk or tk.endswith("%"):
                break
            # Decimal-like token is usually unit price/amount/tax in wrapped tails.
            if re.fullmatch(r"-?\d+\.\d+", tk):
                break
            kept.append(tk)
        return "".join(kept)

    def _parse_item_text(raw_item: str, is_multiline: bool, suffix_text: str = "") -> None:
        line = re.sub(r"\s+", " ", raw_item).strip()
        if not line:
            return
        m = pat_full.match(line) or pat_full_any.search(line)
        qty = "1"
        unit = None
        left = line
        amount = None
        tax = None

        if m:
            left = m.group("left").strip()
            unit = m.group(2)
            qty = m.group("qty")
            amount = m.group("amount")
            tax = m.group("tax")
        else:
            m2 = pat_no_qty.match(line) or pat_no_qty_any.search(line)
            if m2:
                left = m2.group("left").strip()
                unit = m2.group(2)
                amount = m2.group("amount")
                tax = m2.group("tax")
            else:
                md = pat_discount.match(line)
                if md:
                    left = md.group("left").strip()
                    amount = md.group("amount")
                    tax = md.group("tax")
                else:
                    mn = pat_no_unit.match(line) or pat_no_unit_any.search(line)
                    if mn:
                        left = mn.group("left").strip()
                        qty = mn.group("qty")
                        amount = mn.group("amount")
                        tax = mn.group("tax")
                    else:
                        mc = pat_concat_qty_price.match(line)
                        if mc:
                            left = mc.group("left").strip()
                            amount = mc.group("amount")
                            tax = mc.group("tax")
                            qprice = mc.group("qprice")
                            # Split concatenated "<qty><unit_price>" by best fit to amount.
                            best_qty = "1"
                            try:
                                amount_f = float(amount)
                                cand = []
                                for k in (1, 2):
                                    if len(qprice) <= k:
                                        continue
                                    q_s = qprice[:k]
                                    up_s = qprice[k:]
                                    if not q_s.isdigit():
                                        continue
                                    q_i = int(q_s)
                                    if q_i <= 0:
                                        continue
                                    try:
                                        up_f = float(up_s)
                                    except Exception:
                                        continue
                                    if up_f <= 0:
                                        continue
                                    diff = abs(q_i * up_f - amount_f)
                                    cand.append((diff, q_i))
                                if cand:
                                    cand.sort(key=lambda x: x[0])
                                    best_qty = str(cand[0][1])
                            except Exception:
                                best_qty = "1"
                            qty = best_qty

        parts = left.split()
        if not parts:
            return

        if is_multiline:
            # For wrapped item names, keep full merged name in 商品名称.
            merged_name = f"{left}{_clean_suffix_text(suffix_text)}"
            product_line = re.sub(r"\s+", "", merged_name)
            specification = ""
        else:
            product_line = parts[0]
            specification = " ".join(parts[1:]).strip() if len(parts) > 1 else ""

        items.append(
            {
                "product_line": product_line,
                "specification": specification,
                "quantity": qty,
                "unit": unit or "",
                "amount": amount,
                "tax": tax,
            }
        )

    collecting: list[str] = []
    for line in lines:
        if line.startswith("*"):
            if collecting:
                first = collecting[0]
                suffix = "".join(collecting[1:]) if len(collecting) > 1 else ""
                _parse_item_text(first, is_multiline=(len(collecting) > 1), suffix_text=suffix)
            collecting = [line]
            continue

        if collecting:
            if line.startswith(stop_prefixes):
                first = collecting[0]
                suffix = "".join(collecting[1:]) if len(collecting) > 1 else ""
                _parse_item_text(first, is_multiline=(len(collecting) > 1), suffix_text=suffix)
                collecting = []
                continue
            # Skip standalone numeric fragments that are usually broken column wraps.
            if re.fullmatch(r"\d+", line):
                continue
            # Continuation line of current item (wrapped name/spec text).
            collecting.append(line)

    if collecting:
        first = collecting[0]
        suffix = "".join(collecting[1:]) if len(collecting) > 1 else ""
        _parse_item_text(first, is_multiline=(len(collecting) > 1), suffix_text=suffix)

    return items


def _extract_qty_unit(text: str) -> tuple[Optional[str], Optional[str]]:
    qty = unit = None
    fallback_unit = None

    units = r"(个|片|件|盒|瓶|米|套|袋|台|只|块|箱|批)"
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    def _valid_qty(raw: str) -> bool:
        # Reject obvious unit-price-like values such as 1265.486725...
        if "." in raw and len(raw.split(".", 1)[1]) > 3:
            return False
        try:
            val = float(raw)
            if val <= 0:
                return False
            if val > 1000:
                return False
        except ValueError:
            return False
        return True

    # Prefer product detail lines and support both "单位 数量" and "数量 单位".
    for line in lines:
        if not line.startswith("*"):
            continue
        m = re.search(rf"(?:^|\s){units}\s+(\d+(?:\.\d+)?)(?!\*)\b", line)
        if m:
            if _valid_qty(m.group(2)):
                return m.group(2), m.group(1)
            fallback_unit = fallback_unit or m.group(1)
        m = re.search(rf"(?:^|\s)(\d+(?:\.\d+)?)\s*{units}\b", line)
        if m:
            if _valid_qty(m.group(1)):
                return m.group(1), m.group(2)
            fallback_unit = fallback_unit or m.group(2)

    # General fallback in full text.
    m = re.search(rf"(?:^|\s){units}\s+(\d+(?:\.\d+)?)(?!\*)\b", text)
    if m:
        if _valid_qty(m.group(2)):
            return m.group(2), m.group(1)
        fallback_unit = fallback_unit or m.group(1)
    m = re.search(rf"(?:^|\s)(\d+(?:\.\d+)?)\s*{units}\b", text)
    if m:
        if _valid_qty(m.group(1)):
            return m.group(1), m.group(2)
        fallback_unit = fallback_unit or m.group(2)

    return qty, unit or fallback_unit


def _extract_vendor(text: str) -> Optional[str]:
    def _is_placeholder(name: str) -> bool:
        s = re.sub(r"\s+", "", name or "")
        return s in {"", "方", "方方", "买方", "销方", "购方", "售方"}

    def _pick_company_like(s: str) -> Optional[str]:
        candidates = re.findall(
            r"[A-Za-z0-9\u4e00-\u9fa5（）()·\-]{2,}"
            r"(?:股份有限公司|有限责任公司|有限公司|商行|销售中心|中心|制品厂|经营部|工作室|事务所|工厂|厂|店)"
            r"(?:（个体工商户）)?",
            s,
        )
        return candidates[-1].strip() if candidates else None

    patterns = [
        r"(?:销|售)\s*名\s*称\s*[:：]?\s*([^\n\r]*)",
        r"收\s*款\s*方\s*名\s*称\s*[:：]?\s*([^\n\r]*)",
    ]
    for p in patterns:
        m = re.search(p, text)
        if not m:
            continue
        raw = m.group(1).strip()
        company = _pick_company_like(raw) or raw
        company = company.strip()
        if company and not _is_placeholder(company):
            return company

    # Fallback: scan all lines and choose the first company-like seller.
    for line in text.splitlines():
        company = _pick_company_like(line)
        if company and not _is_placeholder(company):
            return company

    return None


def _extract_invoice_number(text: str) -> Optional[str]:
    # 1) Normal form: "发票号码：123..."
    m = re.search(r"发\s*票\s*号\s*码\s*[:：]?\s*([0-9]{8,})", text, flags=re.S)
    if m:
        return m.group(1).strip()

    # 2) Label and value split across lines:
    #    发票号码：
    #    123456...
    lines = [ln.strip() for ln in text.splitlines()]
    for i, ln in enumerate(lines):
        if re.search(r"发\s*票\s*号\s*码\s*[:：]?\s*$", ln):
            for j in range(i + 1, min(i + 6, len(lines))):
                mm = re.search(r"\b([0-9]{8,})\b", lines[j])
                if mm:
                    return mm.group(1)

    # 3) Last fallback: choose first 20-digit-like number in header area.
    head = "\n".join(lines[:25])
    m3 = re.search(r"\b([0-9]{20})\b", head)
    return m3.group(1) if m3 else None


def _item_name_from_product_line(product_line: Optional[str]) -> str:
    if not product_line:
        return "未识别"
    token = product_line.strip().split()[0]
    # Convert "*分类*商品名" -> "商品名"
    if token.startswith("*") and "*" in token[1:]:
        pos = token.find("*", 1)
        if pos != -1:
            tail = token[pos + 1 :].strip()
            if tail:
                return tail
    return token


def extract_invoice_data(pdf_path: str) -> Dict[str, Optional[str]]:
    """Parse invoice PDF and return data dictionary."""
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(_clean_text(page.extract_text() or "") for page in pdf.pages)

    data: Dict[str, Optional[str]] = {
        "vendor": None,
        "invoice_no": None,
        "product_line": None,
        "specification": None,
        "amount": None,
        "tax": None,
        "total": None,
        "quantity": None,
        "unit": None,
    }

    data["vendor"] = _extract_vendor(text)
    data["invoice_no"] = _extract_invoice_number(text)
    amount, tax, total = _extract_money_by_keywords(text)
    data["amount"], data["tax"], data["total"] = amount, tax, total

    items = _extract_invoice_items(text)
    if items:
        first = items[0]
        data["product_line"] = first.get("product_line")
        data["specification"] = first.get("specification")
        data["amount"] = first.get("amount") or data["amount"]
        data["tax"] = first.get("tax") or data["tax"]
        data["quantity"] = first.get("quantity")
        data["unit"] = first.get("unit")
    else:
        product = _extract_product_line(text)
        data.update(product)
        data["quantity"], data["unit"] = _extract_qty_unit(text)

    # attach expanded detail rows for downstream merged writing
    data["items"] = items if items else [  # type: ignore[assignment]
        {
            "product_line": data.get("product_line"),
            "specification": data.get("specification"),
            "quantity": data.get("quantity"),
            "unit": data.get("unit"),
            "amount": data.get("amount"),
            "tax": data.get("tax"),
        }
    ]  # type: ignore[assignment]

    return data


def brand_from_vendor(vendor: str) -> str:
    """Generate a short brand code from vendor name."""
    if not vendor:
        return ""
    try:
        import pypinyin  # type: ignore

        initials = pypinyin.pinyin(vendor, style=pypinyin.Style.FIRST_LETTER)
        letters = [ch for group in initials for ch in group if ch.isalpha()]
        if letters:
            return "".join(letters[:2]).lower()
    except Exception:
        pass

    # fallback
    letters = [c for c in vendor if c.isalpha()]
    return "".join(letters[:2]).lower() if letters else vendor[:2].lower()


def _header_norm(header: Optional[str]) -> str:
    s = str(header or "")
    s = s.replace("\n", "")
    s = re.sub(r"\s+", "", s)
    s = s.replace("（", "(").replace("）", ")")
    return s


def _build_row_by_headers(ws: openpyxl.worksheet.worksheet.Worksheet, data: Dict[str, Optional[str]]) -> dict[int, object]:
    """Map data to columns by header text, compatible with old/new templates."""
    col_values: dict[int, object] = {}
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    for col, raw in enumerate(headers, start=1):
        h = _header_norm(raw)
        v: object = ""

        if "产品分类" in h:
            v = "耗材"
        elif "商品名称" in h:
            v = data.get("product_line") or ""
        elif "CAS号" in h:
            v = ""
        elif "规格" in h:
            v = data.get("specification") or ""
        elif "数量" in h:
            v = _to_int(data.get("quantity"))
        elif "包装单位" in h:
            v = _normalize_unit(data.get("unit"))
        elif "金额" in h and "税额" not in h:
            v = _to_float(data.get("amount"))
        elif "税额" in h:
            v = _to_float(data.get("tax"))
        elif "单价" in h:
            v = None
        elif "品牌" in h:
            v = brand_from_vendor(data.get("vendor") or "")
        elif "货号" in h:
            v = 1
        elif "容量单位" in h:
            v = ""
        elif "容量数字" in h or h == "容量":
            v = ""
        elif "形态" in h:
            v = ""
        elif "特殊购买情况说明" in h:
            v = ""

        col_values[col] = v

    return col_values


def write_to_excel(template_path: str, output_path: str, data: Dict[str, Optional[str]]) -> None:
    """Write a single invoice row into template and save as output_path."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb["导入"]
    row_map = _build_row_by_headers(ws, data)
    for col_index in range(1, ws.max_column + 1):
        ws.cell(row=2, column=col_index, value=row_map.get(col_index, ""))

    # keep only header + one data row in template import sheet
    max_row = ws.max_row
    if max_row > 2:
        ws.delete_rows(3, max_row - 2)

    output_parent = Path(output_path).parent
    output_parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_vendor_excel(template_path: str, output_path: str, rows: list[Dict[str, Optional[str]]]) -> None:
    """Write merged rows for one vendor into one excel."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb["导入"]
    template_row = 2

    for idx, data in enumerate(rows, start=2):
        row_map = _build_row_by_headers(ws, data)
        for col_index in range(1, ws.max_column + 1):
            target = ws.cell(row=idx, column=col_index)
            target.value = row_map.get(col_index, "")

            # Keep template visual/validation format on expanded rows.
            if idx != template_row:
                src = ws.cell(row=template_row, column=col_index)
                target.font = copy(src.font)
                target.fill = copy(src.fill)
                target.border = copy(src.border)
                target.alignment = copy(src.alignment)
                target.number_format = src.number_format
                target.protection = copy(src.protection)

                src_coord = src.coordinate
                for dv in ws.data_validations.dataValidation:
                    if src_coord in dv.cells:
                        dv.add(target.coordinate)

    max_row = ws.max_row
    keep_rows = 1 + len(rows)
    if max_row > keep_rows:
        ws.delete_rows(keep_rows + 1, max_row - keep_rows)

    output_parent = Path(output_path).parent
    output_parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _prepare_output_paths(
    output_root: Path,
    vendor: Optional[str],
    total: Optional[str],
    include_pdf: bool,
    base_pdf: Path,
) -> Dict[str, str]:
    vendor_name = _safe_filename(vendor or "未识别供应商")
    vendor_dir = output_root / vendor_name

    amount_tag = _normalize_money(total)
    base = f"{vendor_name}_{amount_tag}" if amount_tag else vendor_name

    return {
        "vendor_dir": str(vendor_dir),
        "excel_path": str(vendor_dir / f"{base}.xlsx"),
        "pdf_path": str(vendor_dir / f"{base}.pdf") if include_pdf else "",
        "base": base,
        "vendor_name": vendor_name,
    }


def copy_renamed_pdf(src_pdf: str, dst_pdf: str) -> str:
    """Copy PDF and return destination path. Avoid overwrite by suffixing index."""
    src = Path(src_pdf)
    dst = Path(dst_pdf)
    dst.parent.mkdir(parents=True, exist_ok=True)

    if not dst.exists():
        shutil.copy2(src, dst)
        return str(dst)

    stem, suffix = dst.stem, dst.suffix
    idx = 1
    while True:
        candidate = dst.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            shutil.copy2(src, candidate)
            return str(candidate)
        idx += 1


def rename_pdf_in_place(pdf_path: str, vendor: str, total: str) -> str:
    """Rename source PDF in place (backward compatible helper)."""
    amount = _normalize_money(total)
    new_name = f"{_safe_filename(vendor)}_{amount}.pdf" if amount else f"{_safe_filename(vendor)}.pdf"
    new_path = Path(pdf_path).with_name(new_name)
    os.rename(pdf_path, new_path)
    return str(new_path)


def write_summary_excel(output_root: Path, rows: list[dict], summary_name: str = "A物品清单.xlsx") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["物品名称", "公司名称", "金额", "发票号", "总金额"])

    grand_total = 0.0

    for item in rows:
        data = item.get("data", {})
        item_name = _item_name_from_product_line(data.get("product_line"))
        vendor = data.get("vendor") or "未识别供应商"
        total = _to_float(data.get("total"))
        invoice_no = data.get("invoice_no") or ""
        ws.append([item_name, vendor, total, invoice_no, ""])
        grand_total += total

    ws.append(["合计", "", "", "", grand_total])

    # Auto-fit column widths based on content length.
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            txt = "" if val is None else str(val)
            if len(txt) > max_len:
                max_len = len(txt)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 60))

    output_root.mkdir(parents=True, exist_ok=True)
    out = output_root / summary_name
    wb.save(out)
    return str(out)


def _batch_process(args: argparse.Namespace) -> list[dict]:
    pdf_root = Path(args.input_dir)
    pdf_files = sorted(pdf_root.rglob("*.pdf")) if args.recursive else sorted(pdf_root.glob("*.pdf"))
    results = []
    vendor_rows: dict[str, list[Dict[str, Optional[str]]]] = defaultdict(list)
    vendor_dir_map: dict[str, str] = {}
    seen_invoice_nos: set[str] = set()

    for pdf_file in pdf_files:
        if pdf_file.name == ".DS_Store":
            continue

        data = extract_invoice_data(str(pdf_file))
        invoice_no = (data.get("invoice_no") or "").strip()
        if invoice_no:
            if invoice_no in seen_invoice_nos:
                print(f"[SKIP] duplicate invoice_no={invoice_no}: {pdf_file}")
                continue
            seen_invoice_nos.add(invoice_no)
        paths = _prepare_output_paths(
            output_root=Path(args.output_dir),
            vendor=data.get("vendor"),
            total=data.get("total"),
            include_pdf=True,
            base_pdf=pdf_file,
        )

        vendor_key = paths["vendor_name"]
        expanded_items = data.get("items") or []
        if isinstance(expanded_items, list) and expanded_items:
            for item in expanded_items:
                merged = dict(data)
                if isinstance(item, dict):
                    if "product_line" in item:
                        merged["product_line"] = item.get("product_line")
                    if "specification" in item:
                        merged["specification"] = item.get("specification")
                    if "quantity" in item:
                        merged["quantity"] = item.get("quantity")
                    if "unit" in item:
                        merged["unit"] = item.get("unit")
                    if "amount" in item:
                        merged["amount"] = item.get("amount")
                    if "tax" in item:
                        merged["tax"] = item.get("tax")
                vendor_rows[vendor_key].append(merged)
        else:
            vendor_rows[vendor_key].append(data)
        vendor_dir_map[vendor_key] = paths["vendor_dir"]

        pdf_out = ""
        if args.rename:
            pdf_out = copy_renamed_pdf(str(pdf_file), paths["pdf_path"])
            print(f"[OK] PDF : {pdf_out}")

        results.append({"pdf": str(pdf_file), "excel": "", "renamed_pdf": pdf_out or "", "data": data})

    for vendor_key, rows in vendor_rows.items():
        vendor_dir = Path(vendor_dir_map[vendor_key])
        excel_path = str(vendor_dir / f"{vendor_key}.xlsx")
        write_vendor_excel(args.template, excel_path, rows)
        print(f"[OK] Excel: {excel_path}")
        for item in results:
            if (item.get("data") or {}).get("vendor") == vendor_key:
                item["excel"] = excel_path

    if not args.no_summary:
        summary_path = write_summary_excel(Path(args.output_dir), results, args.summary_name)
        print(f"[OK] Summary: {summary_path}")

    return results


def _single_process(args: argparse.Namespace) -> list[dict]:
    data = extract_invoice_data(args.pdf)

    out_path = args.output
    if not out_path:
        output_root = Path(args.output_dir)
        paths = _prepare_output_paths(
            output_root=output_root,
            vendor=data.get("vendor"),
            total=data.get("total"),
            include_pdf=False,
            base_pdf=Path(args.pdf),
        )
        out_path = paths["excel_path"]

    write_to_excel(args.template, out_path, data)
    print(f"[OK] Excel: {out_path}")

    renamed_pdf = ""
    if args.rename:
        if data.get("vendor") and data.get("total"):
            renamed_pdf = rename_pdf_in_place(args.pdf, data["vendor"], data["total"])
            print(f"[OK] PDF renamed: {renamed_pdf}")
        else:
            print("[WARN] skip rename: missing vendor or total")

    return [{"pdf": args.pdf, "excel": out_path, "renamed_pdf": renamed_pdf, "data": data}]


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse VAT invoice PDF(s) and generate Excel file(s).")
    parser.add_argument("--pdf", help="Single invoice PDF path")
    parser.add_argument("--input-dir", default=".", help="Directory containing multiple invoice PDFs")
    parser.add_argument("--template", default="模板文件.xlsx", help="Path to the Excel template")
    parser.add_argument("--output", default="", help="Output path for single PDF mode")
    parser.add_argument("--output-dir", default="输出结果", help="Output root for batch mode")
    parser.add_argument("--rename", action="store_true", help="Rename/copy PDF to <vendor>_<total>.pdf")
    parser.add_argument("--recursive", action="store_true", help="Search pdfs recursively under input-dir")
    parser.add_argument("--summary-name", default="A物品清单.xlsx", help="Summary excel filename in output root")
    parser.add_argument("--no-summary", action="store_true", help="Do not generate summary excel in batch mode")

    args = parser.parse_args()

    if not args.pdf and not os.path.isdir(args.input_dir):
        raise ValueError("Need either --pdf or valid --input-dir")

    if args.pdf and not os.path.exists(args.pdf):
        raise FileNotFoundError(f"PDF not found: {args.pdf}")

    if args.pdf:
        _single_process(args)
        return

    _batch_process(args)


if __name__ == "__main__":
    main()
